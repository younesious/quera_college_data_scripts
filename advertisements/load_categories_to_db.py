import pymysql
from openpyxl import load_workbook

conn = pymysql.connect(host='localhost', user='root', passwd='pass', db='quera_college')
cursor = conn.cursor()

# you can replace Category.xlsx with your filename
workbook = load_workbook('Category.xlsx')
sheet = workbook.active

columns = []
for i in range(1, sheet.max_column + 1):
    columns.append(sheet.cell(row=1, column=i).value)

table_name = 'categories'

table = f'DROP TABLE IF EXISTS {table_name};'
cursor.execute(table)

schema = f'CREATE TABLE {table_name} (`ID` BIGINT UNSIGNED PRIMARY KEY,'
for i, col in enumerate(columns):
    if col == 'ID':
        continue
    datatype = 'varchar(255)' if isinstance(col, str) else 'float'
    datatype += ' NULL'
    schema += f"`{col}` {datatype}"
    if i < len(columns) - 1:
        schema += ', '
schema += ');'

cursor.execute(schema)

for i in range(2, sheet.max_row + 1):
    row_data = []
    for j in range(1, sheet.max_column + 1):
        row_data.append(sheet.cell(row=i, column=j).value)
    values = str(tuple(row_data)).replace('None', 'null')	
    insert_stmt = f"INSERT INTO {table_name} VALUES {values};"
    cursor.execute(insert_stmt)

conn.commit()
conn.close()

print("Excel file imported to MySQL database!")

